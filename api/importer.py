"""
Katalog + ombor qoldiqlari IMPORTI (TZ P0-4).

Qabul qilish mezonlari:
    * Standart shablon: nomi, xususiyatlari, o'lchov birligi, qoldiq, tannarx
    * Format xatolari QATOR BO'YICHA ajratib ko'rsatiladi

DIZAYN TAMOYILLARI
------------------
1. XATO BITTA QATORNI TO'XTATADI, IMPORTNI EMAS. Har bir muammoli qator
   `errors` ro'yxatiga tushadi (qator raqami + ustun + o'zbekcha xabar), qolgan
   qatorlar normal import bo'ladi. Natija: "N ta qator qabul qilindi, M ta
   qatorda xato".
2. QATOR RAQAMI — FAYLDAGI HAQIQIY RAQAM. Foydalanuvchi Excel'da ko'radigan
   raqam bilan bir xil (sarlavha qatori ham sanaladi), aks holda xatoni topib
   bo'lmaydi.
3. TAXMIN QILINMAYDI. Sonni ishonchli o'qib bo'lmasa — xato, "0" emas.
   Kategoriya topilmasa — ogohlantirish, mahsulot kategoriyasiz qo'shiladi.
4. DRY-RUN bazaga UMUMAN tegmaydi (partiya yozuvi ham yaratilmaydi).

QO'LLAB-QUVVATLANADIGAN FORMATLAR
---------------------------------
    .xlsx  — openpyxl (birinchi varaq)
    .csv   — utf-8 / utf-8-sig / cp1251, ajratgich , ; yoki TAB (avtomatik)
    Google Sheets — "Fayl > Yuklab olish > CSV" orqali (alohida integratsiya
    shart emas: eksport qilingan CSV shu yerda o'qiladi).
"""
import csv
import io
import re
import unicodedata
import uuid
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

from api import db

# ---------------------------------------------------------------------------
# 1. Ustun sarlavhalarini tanish
# ---------------------------------------------------------------------------
# Uch tilda (o'zbek lotin/kirill, rus, ingliz). Katta-kichik harf, apostrof
# turi, qavs ichidagi izoh va tinish belgilari e'tiborga olinmaydi.
# TARTIB MUHIM: aniqroq atama (tannarx) umumiyroqdan (narx) OLDIN turishi shart
# emas — moslashtirish eng UZUN alias bo'yicha amalga oshadi (_match_alias).
COLUMN_ALIASES: Dict[str, List[str]] = {
    "name": [
        "nomi", "nom", "nomlanishi", "mahsulot", "mahsulot nomi", "tovar",
        "tovar nomi", "mahsulot/xizmat", "xizmat nomi",
        "наименование", "название", "товар", "наименование товара", "продукт",
        "name", "product", "product name", "item", "item name", "title",
    ],
    "keywords": [
        "xususiyatlari", "xususiyati", "xususiyat", "xossalari", "tavsifi",
        "tavsif", "kalit so'zlar", "kalit so'z", "izoh",
        "характеристики", "характеристика", "свойства", "описание",
        "ключевые слова", "примечание",
        "properties", "attributes", "features", "keywords", "description",
        "specs", "spec",
    ],
    "unit": [
        "o'lchov birligi", "olchov birligi", "o'lchov", "birligi", "birlik",
        "единица измерения", "ед. изм.", "ед изм", "ед.изм", "единица",
        "unit", "uom", "measure", "measurement unit",
    ],
    "stock_qty": [
        "qoldiq", "qoldig'i", "ombor qoldig'i", "ombordagi qoldiq", "zaxira",
        "miqdor", "miqdori", "soni", "mavjud",
        "остаток", "остатки", "остаток на складе", "количество", "кол-во",
        "наличие", "запас",
        "stock", "qty", "quantity", "balance", "on hand", "in stock",
        "stock qty", "available",
    ],
    "cost_price": [
        "tannarx", "tannarxi", "tan narx", "xarid narxi", "kirim narxi",
        "себестоимость", "закупочная цена", "цена закупки", "закуп",
        "cost", "cost price", "unit cost", "purchase price",
    ],
    "price": [
        "narx", "narxi", "sotuv narxi", "sotish narxi",
        "цена", "цена продажи", "продажная цена", "стоимость",
        "price", "sale price", "selling price",
    ],
    "category_code": [
        "kategoriya", "kategoriyasi", "toifa",
        "категория", "категории",
        "category", "category code",
    ],
    "currency": [
        "valyuta", "valyutasi",
        "валюта",
        "currency", "cur",
    ],
}

# Shablonda va xato xabarlarida ko'rinadigan o'zbekcha nomlar
FIELD_LABELS: Dict[str, str] = {
    "name": "Nomi",
    "keywords": "Xususiyatlari",
    "unit": "O‘lchov birligi",
    "stock_qty": "Qoldiq",
    "cost_price": "Tannarx",
    "price": "Sotuv narxi",
    "category_code": "Kategoriya",
    "currency": "Valyuta",
}

REQUIRED_FIELDS = ("name",)          # busiz import mumkin emas
TEMPLATE_FIELDS = ("name", "keywords", "unit", "stock_qty", "cost_price",
                   "price", "category_code", "currency")

_APOSTROPHES = "‘’ʻʼ`´"   # ‘ ’ ʻ ʼ ` ´
_SPACES = "    "                     # NBSP va boshqa probellar


def _norm_header(s: Any) -> str:
    """Sarlavhani solishtirishga tayyorlaydi: kichik harf, apostroflarni
    yagona ' ga keltirish, qavs ichini olib tashlash, ortiqcha belgilarni
    tozalash. "Tannarx (UZS)*" -> "tannarx"."""
    if s is None:
        return ""
    t = unicodedata.normalize("NFKC", str(s)).strip().lower()
    for ch in _APOSTROPHES:
        t = t.replace(ch, "'")
    for ch in _SPACES:
        t = t.replace(ch, " ")
    t = re.sub(r"\([^)]*\)", " ", t)     # qavs ichidagi izoh
    t = re.sub(r"[*:;#№]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    t = t.strip(" .,-_")
    return t


# Eng uzun alias birinchi — "tannarx" "narx" dan ustun bo'lishi uchun
_ALIAS_INDEX: List[Tuple[str, str]] = sorted(
    ((_norm_header(a), f) for f, al in COLUMN_ALIASES.items() for a in al),
    key=lambda x: -len(x[0]),
)


def _match_alias(header: str) -> Optional[str]:
    """Sarlavha -> maydon nomi. Avval aniq moslik, keyin ichiga kirish
    (eng uzun alias g'olib: "Tannarx, so'm" -> cost_price, "narx" emas)."""
    h = _norm_header(header)
    if not h:
        return None
    for alias, field in _ALIAS_INDEX:
        if h == alias:
            return field
    for alias, field in _ALIAS_INDEX:
        if len(alias) >= 3 and alias in h:
            return field
    return None


def detect_columns(headers: List[Any]) -> Tuple[Dict[str, int], List[str]]:
    """Sarlavha qatoridan {maydon: ustun indeksi} va tanilmagan sarlavhalar."""
    mapping: Dict[str, int] = {}
    unknown: List[str] = []
    for idx, h in enumerate(headers):
        if h is None or str(h).strip() == "":
            continue
        field = _match_alias(h)
        if field and field not in mapping:
            mapping[field] = idx
        elif field is None:
            unknown.append(str(h).strip())
    return mapping, unknown


# ---------------------------------------------------------------------------
# 2. Son o'qish (import uchun)
# ---------------------------------------------------------------------------
_BLANKS = {"", "-", "—", "–", "n/a", "na", "yo'q", "нет", "null", "none"}


def parse_number(raw: Any) -> Tuple[Optional[Decimal], Optional[str]]:
    """Matn/katak qiymatidan sonni ajratadi.

    Qaytaradi: (qiymat, xato). Bo'sh katak -> (None, None) — bu xato emas.
    Ishonchli o'qib bo'lmasa -> (None, "sabab") va qator xatoga tushadi.

    Qoidalar (chalkash formatlar uchun aniq va oldindan aytiladigan):
        "1 234"      -> 1234    (probel — mingliklar ajratgichi, 3 xonali guruh)
        "1 234,56"   -> 1234.56 (oxirgi ajratgich — kasr)
        "1,234.56"   -> 1234.56
        "12,5"       -> 12.5    (yolg'iz vergul — kasr)
        "1,234"      -> 1234    (aniq mingliklar naqshi: 1-3 xona + ,000)
        "120 dona"   -> 120     (birlik qo'shimchasi tashlanadi)
        "10-15"      -> xato    (diapazon — taxmin qilmaymiz)
    """
    if raw is None:
        return None, None
    if isinstance(raw, bool):
        return None, "mantiqiy qiymat (HA/YO‘Q) son o‘rniga kelgan"
    if isinstance(raw, (int, float, Decimal)):
        try:
            return Decimal(str(raw)), None
        except InvalidOperation:
            return None, f"son sifatida o‘qib bo‘lmadi: {raw!r}"

    s = unicodedata.normalize("NFKC", str(raw)).strip()
    for ch in _SPACES:
        s = s.replace(ch, " ")
    if s.lower() in _BLANKS:
        return None, None

    m = re.match(r"^([+-]?\d[\d \.,]*)", s)
    if not m:
        return None, f"son emas: “{s}”"
    token = m.group(1).strip()
    tail = s[m.end():]
    if re.search(r"\d", tail):
        # "10-15", "2 x 3", "5 dona 3 quti" — qaysi son ekani noma'lum
        return None, f"bir nechta son yoki diapazon: “{s}” (bitta son kiriting)"

    sign = "-" if token.startswith("-") else ""
    token = token.lstrip("+-")

    # Probel — faqat to'g'ri mingliklar guruhida ruxsat (1 234 567,80)
    if " " in token:
        if not re.fullmatch(r"\d{1,3}( \d{3})+([.,]\d+)?", token):
            return None, f"sonda ortiqcha probel: “{s}”"
        token = token.replace(" ", "")

    has_dot, has_comma = "." in token, "," in token
    if has_dot and has_comma:
        # Oxirgi ajratgich — kasr, ikkinchisi mingliklar
        dec = "," if token.rfind(",") > token.rfind(".") else "."
        token = token.replace("." if dec == "," else ",", "").replace(dec, ".")
    elif has_comma:
        token = (token.replace(",", "")
                 if re.fullmatch(r"\d{1,3}(,\d{3})+", token)
                 else token.replace(",", "."))
    elif has_dot and token.count(".") > 1:
        token = token.replace(".", "")      # 1.234.567 — mingliklar

    try:
        return Decimal(sign + token), None
    except InvalidOperation:
        return None, f"son sifatida o‘qib bo‘lmadi: “{s}”"


def _split_keywords(raw: Any) -> List[str]:
    """"монитор; стол, 24 dyum" -> ['монитор', 'стол', '24 dyum'] (takrorsiz)."""
    if raw is None:
        return []
    parts = re.split(r"[;|\n,]+", str(raw))
    out: List[str] = []
    for p in parts:
        v = p.strip(" .\t")
        if v and v not in out:
            out.append(v)
    return out[:30]


#: PostgreSQL TEXT ustuni NUL baytini (\x00) SAQLAY OLMAYDI — insert
#: DataError bilan uziladi. Boshqa boshqaruv belgilari (\x01-\x08 va h.k.)
#: matnni ko'rinmas holda buzadi. Ikkalasi ham Excel'dan eksport qilingan
#: fayllarda uchraydi, shuning uchun bitta joyda — hamma matn shu yerdan o'tadi.
_NAZORAT_BELGI = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def _cell_text(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    v = _NAZORAT_BELGI.sub("", str(raw)).strip()
    return v or None


# ---------------------------------------------------------------------------
# 3. Fayl o'qish (.xlsx / .csv)
# ---------------------------------------------------------------------------
class ImportFormatError(ValueError):
    """Butun faylga tegishli xato (qatorga emas) — import boshlanmaydi."""


def _read_xlsx(data: bytes) -> List[List[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:      # pragma: no cover
        raise ImportFormatError(
            "Excel o‘qish uchun `openpyxl` o‘rnatilmagan.") from e
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise ImportFormatError(
            f"Excel fayli ochilmadi (buzilgan yoki .xlsx emas): {e}") from e
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _read_csv(data: bytes) -> List[List[Any]]:
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1251", "cp1252"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ImportFormatError(
            "CSV kodlashi aniqlanmadi. Faylni UTF-8 da saqlang.")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        delim = dialect.delimiter
    except csv.Error:
        # Sniffer aniqlay olmasa — birinchi qatorda qaysi belgi ko'proq
        head = sample.splitlines()[0] if sample.splitlines() else ""
        delim = max(",;\t", key=head.count)
    return [row for row in csv.reader(io.StringIO(text), delimiter=delim)]


def read_table(data: bytes, filename: str) -> Tuple[List[List[Any]], str]:
    """Fayl baytlaridan xom jadval + format nomi. Kengaytma bo'yicha tanlaydi,
    noaniq bo'lsa ZIP imzosiga (.xlsx = PK) qaraydi."""
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return _read_xlsx(data), "xlsx"
    if name.endswith((".csv", ".txt", ".tsv")):
        return _read_csv(data), "csv"
    if name.endswith((".xls", ".xlsm", ".ods")):
        raise ImportFormatError(
            "Bu format qo‘llab-quvvatlanmaydi. Faylni .xlsx yoki .csv "
            "ko‘rinishida saqlang (Excel: “Farqli saqlash > .xlsx”).")
    if data[:2] == b"PK":
        return _read_xlsx(data), "xlsx"
    return _read_csv(data), "csv"


def _find_header(rows: List[List[Any]], limit: int = 10) -> Tuple[int, Dict[str, int], List[str]]:
    """Sarlavha qatorini topadi (fayl boshida sarlavha/izoh qatorlari bo'lishi
    mumkin). Shart: `name` ustuni tanilgan bo'lsin."""
    best: Optional[Tuple[int, Dict[str, int], List[str]]] = None
    for i, row in enumerate(rows[:limit]):
        if not row:
            continue
        mapping, unknown = detect_columns(row)
        if "name" in mapping:
            if best is None or len(mapping) > len(best[1]):
                best = (i, mapping, unknown)
    if best is None:
        raise ImportFormatError(
            "Sarlavha qatori topilmadi: “Nomi” (yoki Наименование / Name) "
            "ustuni yo‘q. Namunaviy shablonni yuklab oling.")
    return best


# ---------------------------------------------------------------------------
# 4. Qatorlarni tekshirish
# ---------------------------------------------------------------------------
def _err(row_no: int, field: Optional[str], value: Any, message: str) -> Dict[str, Any]:
    return {
        "row": row_no,
        "column": FIELD_LABELS.get(field or "", field or "—"),
        "field": field,
        "value": None if value is None else str(value)[:120],
        "message": message,
    }


def parse_rows(rows: List[List[Any]], mapping: Dict[str, int], header_idx: int,
               known_categories: Optional[set] = None
               ) -> Tuple[List[Dict[str, Any]], List[Dict], List[Dict]]:
    """Xom jadvaldan tozalangan yozuvlar + xatolar + ogohlantirishlar.

    `row_no` — FAYLDAGI qator raqami (1 dan, sarlavha ham sanaladi), ya'ni
    foydalanuvchi Excel'da ko'radigan raqam bilan bir xil.
    """
    ok: List[Dict[str, Any]] = []
    errors: List[Dict] = []
    warnings: List[Dict] = []
    seen_names: Dict[str, int] = {}

    def cell(row: List[Any], field: str) -> Any:
        idx = mapping.get(field)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return v if not isinstance(v, str) or v.strip() else None

    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        row_no = i + 1                       # 1-asosli (Excel bilan bir xil)
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue                          # bo'sh qator — jim o'tkazamiz

        row_errors: List[Dict] = []

        # --- Nomi (majburiy) ---
        name = _cell_text(cell(row, "name"))
        if not name:
            row_errors.append(_err(row_no, "name", None,
                                   "Mahsulot nomi bo‘sh — qator qabul qilinmadi."))
        elif len(name) > 500:
            row_errors.append(_err(row_no, "name", name,
                                   "Nom juda uzun (500 belgidan ko‘p)."))
        elif name.lower() in seen_names:
            row_errors.append(_err(
                row_no, "name", name,
                f"Nom faylda takrorlangan ({seen_names[name.lower()]}-qatorda "
                f"ham bor). Qatorlarni birlashtiring."))

        # --- Sonli maydonlar ---
        nums: Dict[str, Optional[Decimal]] = {}
        for field in ("stock_qty", "cost_price", "price"):
            if field not in mapping:
                continue
            val, err = parse_number(cell(row, field))
            if err:
                row_errors.append(_err(row_no, field, cell(row, field),
                                       f"{FIELD_LABELS[field]} — {err}."))
            elif val is not None and val < 0:
                row_errors.append(_err(row_no, field, cell(row, field),
                                       f"{FIELD_LABELS[field]} manfiy bo‘lishi mumkin emas."))
            else:
                nums[field] = val

        # --- Valyuta ---
        currency = _cell_text(cell(row, "currency"))
        if currency:
            currency = currency.upper()[:8]
            if not re.fullmatch(r"[A-Z]{3}", currency):
                row_errors.append(_err(row_no, "currency", currency,
                                       "Valyuta 3 harfli kod bo‘lishi kerak (UZS, USD, EUR)."))
                currency = None

        # --- Kategoriya (xato emas — ogohlantirish) ---
        category = _cell_text(cell(row, "category_code"))
        if category and known_categories is not None and category not in known_categories:
            warnings.append(_err(
                row_no, "category_code", category,
                f"“{category}” kategoriyasi ma’lumotnomada yo‘q — mahsulot "
                f"kategoriyasiz qo‘shildi."))
            category = None

        if row_errors:
            errors.extend(row_errors)
            continue

        seen_names[name.lower()] = row_no

        # Qoldiq ustuni bor, lekin katak bo'sh — import to'xtamaydi, ogohlantirish
        if "stock_qty" in mapping and nums.get("stock_qty") is None:
            warnings.append(_err(row_no, "stock_qty", None,
                                 "Qoldiq ko‘rsatilmagan — mavjud qiymat o‘zgarmaydi."))

        ok.append({
            "row": row_no,
            "name": name,
            "keywords": _split_keywords(cell(row, "keywords")),
            "unit": _cell_text(cell(row, "unit")),
            "stock_qty": nums.get("stock_qty"),
            "cost_price": nums.get("cost_price"),
            "price": nums.get("price"),
            "category_code": category,
            "currency": currency,
        })

    return ok, errors, warnings


# ---------------------------------------------------------------------------
# 5. Bazaga yozish
# ---------------------------------------------------------------------------
_BATCH_INSERT_SQL = """
INSERT INTO catalog_import_batch (id, filename, source, rows_total, rows_ok,
                                  rows_error, inserted, updated, errors)
VALUES (%(id)s, %(filename)s, %(source)s, %(rows_total)s, %(rows_ok)s,
        %(rows_error)s, 0, 0, %(errors)s)
"""

_BATCH_FINISH_SQL = """
UPDATE catalog_import_batch SET inserted=%(inserted)s, updated=%(updated)s
WHERE id=%(id)s
"""

_FIND_SQL = """
SELECT id FROM catalog_product
WHERE lower(name) = lower(%(name)s)
ORDER BY id LIMIT 1
"""

_INSERT_SQL = """
INSERT INTO catalog_product
    (name, category_code, keywords, unit, price, currency, notify,
     stock_qty, stock_unit, stock_updated_at, cost_price, import_batch_id)
VALUES
    (%(name)s, %(category_code)s, %(keywords)s, %(unit)s, %(price)s,
     %(currency)s, TRUE, %(stock_qty)s, %(stock_unit)s,
     CASE WHEN %(stock_qty)s IS NULL THEN NULL ELSE now() END,
     %(cost_price)s, %(batch_id)s)
RETURNING id
"""

# COALESCE emas, aniq shart: faylda ustun BO'LMASA yoki katak BO'SH bo'lsa
# mavjud qiymat SAQLANADI. Import hech qachon ma'lumotni jimgina o'chirmaydi.
_UPDATE_SQL = """
UPDATE catalog_product SET
    -- Nom fayldagi yozilishiga keltiriladi (moslik lower() bo'yicha topilgan):
    -- shunda "NOUTBUK" -> "Noutbuk" kabi tuzatishni import orqali qilish mumkin.
    name             = %(name)s,
    keywords         = CASE WHEN %(has_keywords)s THEN %(keywords)s ELSE keywords END,
    unit             = COALESCE(%(unit)s, unit),
    category_code    = COALESCE(%(category_code)s, category_code),
    currency         = COALESCE(%(currency)s, currency),
    price            = COALESCE(%(price)s, price),
    cost_price       = COALESCE(%(cost_price)s, cost_price),
    stock_qty        = COALESCE(%(stock_qty)s, stock_qty),
    stock_unit       = COALESCE(%(stock_unit)s, stock_unit),
    stock_updated_at = CASE WHEN %(stock_qty)s IS NULL
                            THEN stock_updated_at ELSE now() END,
    import_batch_id  = %(batch_id)s,
    updated_at       = now()
WHERE id = %(id)s
"""


def _known_categories() -> Optional[set]:
    """dim_category_uz kodlari (kategoriya tekshiruvi uchun). Baza yo'q bo'lsa
    None — u holda kategoriya tekshirilmaydi (import baribir ishlaydi)."""
    try:
        rows = db.query("SELECT code FROM dim_category_uz")
        return {r["code"] for r in rows}
    except Exception:
        return None


def import_catalog(data: bytes, filename: str, *, dry_run: bool = True) -> Dict[str, Any]:
    """P0-4 asosiy kirish nuqtasi.

    dry_run=True  — faqat tekshiradi, bazaga HECH NARSA yozilmaydi
                    (partiya yozuvi ham yaratilmaydi).
    dry_run=False — to'g'ri qatorlar bitta tranzaksiyada yoziladi.

    Mavjud mahsulot bilan moslashtirish: NOM bo'yicha, katta-kichik harf
    farqsiz. Topilsa yangilanadi, topilmasa qo'shiladi.
    """
    rows, fmt = read_table(data, filename)
    if not rows:
        raise ImportFormatError("Fayl bo‘sh.")

    header_idx, mapping, unknown = _find_header(rows)
    missing = [FIELD_LABELS[f] for f in REQUIRED_FIELDS if f not in mapping]

    ok, errors, warnings = parse_rows(rows, mapping, header_idx,
                                      known_categories=_known_categories())

    result: Dict[str, Any] = {
        "dry_run": dry_run,
        "filename": filename,
        "format": fmt,
        "batch_id": None,
        "columns": {
            "detected": {FIELD_LABELS[f]: (rows[header_idx][i] if i < len(rows[header_idx]) else None)
                         for f, i in mapping.items()},
            "unknown": unknown,          # tanilmagan sarlavhalar — e'tiborsiz qoldirildi
            "missing": missing,
        },
        "header_row": header_idx + 1,
        "rows_total": len(ok) + len({e["row"] for e in errors}),
        "rows_ok": len(ok),
        "rows_error": len({e["row"] for e in errors}),
        "inserted": 0,
        "updated": 0,
        "errors": errors,
        "warnings": warnings,
        "preview": [_preview(r) for r in ok[:50]],
    }

    if dry_run or not ok:
        # Dry-run: nima bo'lishini OLDINDAN aytamiz (yozmasdan)
        if ok:
            result["inserted"], result["updated"] = _forecast(ok)
        return result

    batch_id = str(uuid.uuid4())
    inserted = updated = 0
    import json as _json
    with db.get_conn() as conn:
        try:
            with conn.cursor() as cur:
                cur.execute(_BATCH_INSERT_SQL, {
                    "id": batch_id, "filename": filename, "source": fmt,
                    "rows_total": result["rows_total"], "rows_ok": result["rows_ok"],
                    "rows_error": result["rows_error"],
                    "errors": _json.dumps(errors, ensure_ascii=False),
                })
                for r in ok:
                    cur.execute(_FIND_SQL, {"name": r["name"]})
                    found = cur.fetchone()
                    params = {
                        "name": r["name"],
                        "category_code": r["category_code"],
                        "keywords": r["keywords"],
                        "has_keywords": bool(r["keywords"]),
                        "unit": r["unit"],
                        "price": r["price"],
                        "cost_price": r["cost_price"],
                        "currency": r["currency"],
                        "stock_qty": r["stock_qty"],
                        # Qoldiq birligi ko'rsatilmasa — sotuv birligi bilan bir xil
                        "stock_unit": r["unit"],
                        "batch_id": batch_id,
                    }
                    if found:
                        cur.execute(_UPDATE_SQL, {**params, "id": found["id"]})
                        updated += 1
                    else:
                        cur.execute(_INSERT_SQL, params)
                        inserted += 1
                cur.execute(_BATCH_FINISH_SQL, {
                    "id": batch_id, "inserted": inserted, "updated": updated})
            conn.commit()
        except Exception:
            conn.rollback()
            raise

    result["batch_id"] = batch_id
    result["inserted"] = inserted
    result["updated"] = updated
    return result


def _forecast(ok: List[Dict[str, Any]]) -> Tuple[int, int]:
    """Dry-run uchun: nechtasi qo'shiladi / nechtasi yangilanadi."""
    try:
        rows = db.query("SELECT lower(name) AS n FROM catalog_product")
    except Exception:
        return len(ok), 0
    existing = {r["n"] for r in rows}
    upd = sum(1 for r in ok if r["name"].lower() in existing)
    return len(ok) - upd, upd


def _preview(r: Dict[str, Any]) -> Dict[str, Any]:
    """Frontend jadvali uchun (Decimal -> float)."""
    return {
        "row": r["row"], "name": r["name"], "keywords": r["keywords"],
        "unit": r["unit"],
        "stock_qty": float(r["stock_qty"]) if r["stock_qty"] is not None else None,
        "cost_price": float(r["cost_price"]) if r["cost_price"] is not None else None,
        "price": float(r["price"]) if r["price"] is not None else None,
        "category_code": r["category_code"], "currency": r["currency"],
    }


# ---------------------------------------------------------------------------
# 6. Namunaviy shablon (GET /catalog/import/template)
# ---------------------------------------------------------------------------
_TEMPLATE_HEADERS = ["Nomi", "Xususiyatlari", "O‘lchov birligi", "Qoldiq",
                     "Tannarx", "Sotuv narxi", "Kategoriya", "Valyuta"]

_TEMPLATE_ROWS = [
    ["Noutbuk Lenovo V15", "15.6 dyum; 8GB RAM; SSD 256GB", "dona", 12,
     6500000, 7800000, "elektronika", "UZS"],
    ["Ofis stoli 120x60", "LDSP; oq rang", "dona", 40, 850000, 1150000, "", "UZS"],
    ["A4 qog‘oz 80 g/m²", "500 varaq; Snegurochka", "quti", 250, 42000, 55000,
     "", "UZS"],
]


def template_xlsx() -> bytes:
    """Namunaviy .xlsx shablon (sarlavhalar + 3 ta misol qator)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Katalog"
    ws.append(_TEMPLATE_HEADERS)
    fill = PatternFill("solid", fgColor="E8EEF7")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = fill
        c.alignment = Alignment(vertical="center")
    for r in _TEMPLATE_ROWS:
        ws.append(r)
    for i, w in enumerate([34, 40, 16, 12, 16, 16, 18, 10], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    # Ikkinchi varaq — qisqa yo'riqnoma (import faqat 1-varaqni o'qiydi)
    info = wb.create_sheet("Yo‘riqnoma")
    for line in [
        ["Ustunlar tartibi muhim EMAS — sarlavha nomi bo‘yicha tanib olinadi."],
        ["Majburiy ustun: Nomi. Qolganlari ixtiyoriy."],
        ["Sarlavhalar o‘zbek, rus yoki ingliz tilida bo‘lishi mumkin."],
        ["Qoldiq — ombordagi mavjud miqdor (masalan 120 yoki 12,5)."],
        ["Tannarx — xarid narxi. Sotuv narxi alohida ustun."],
        ["Xususiyatlari — nuqtali vergul bilan ajrating: 15.6 dyum; 8GB RAM"],
        ["Bo‘sh qoldirilgan katak mavjud qiymatni O‘ZGARTIRMAYDI."],
        ["Nom bo‘yicha moslashtiriladi: bor bo‘lsa yangilanadi, yo‘q bo‘lsa qo‘shiladi."],
        ["Google Sheets: Fayl > Yuklab olish > CSV, so‘ng shu faylni yuklang."],
    ]:
        info.append(line)
    info.column_dimensions["A"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def template_csv() -> bytes:
    """Namunaviy .csv shablon (Excel uchun BOM bilan — kirill buzilmasin)."""
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", lineterminator="\r\n")
    w.writerow(_TEMPLATE_HEADERS)
    for r in _TEMPLATE_ROWS:
        w.writerow(r)
    return buf.getvalue().encode("utf-8-sig")
